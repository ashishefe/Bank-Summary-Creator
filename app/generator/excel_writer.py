"""Generate Bank Summary Excel file matching the CA's exact format.

Output format per bank account section:
  Row: BANK & BRANCH: [bank]  [branch]
  Row: A/C NO.: [number]
  Row: IFSC [code]
  Row: PARTICULARS | TOTAL | APR | MAY | ... | MAR
  Row: Opening Balance. | [value] | [monthly chain]
  Row: Deposits :-
  Row:   [category] | =SUM(C:N) | [monthly amounts]
  Row:   Total Rs. | =SUM(...) | =SUM(...)
  Row: Withdrawls:-
  Row:   [category] | =SUM(C:N) | [monthly amounts]
  Row:   Total Rs. | =SUM(...) | =SUM(...)
  Row: Closing Balance | =Open+Dep-Wdl | per month
  Row: (blank)
  Row: [verification column O with independent value, P = diff check]
"""

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from collections import defaultdict
from typing import Optional

from app.config import FY_MONTHS
from app.generator.summary_builder import AccountSummary
from app.models.domain import BankAccount


def _col(n: int) -> str:
    """Column letter for 0-based index. 0=A, 1=B, 2=C, ..., 13=N."""
    return get_column_letter(n + 1)


def generate_excel(
    summaries: list[AccountSummary],
    client_name: str,
    assessment_year: str = "2025-2026",
    output_path: Path | str = "Bank Summary.xlsx",
    accounts: Optional[list[BankAccount]] = None,
) -> Path:
    """Generate the Bank Summary Excel workbook.

    Uses Excel formulas for totals and closing balances (not computed values).
    If accounts are provided, adds a detail sheet per account with line-by-line
    transactions grouped by category.
    """
    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = f"Bank summary"

    # Styles
    bold = Font(bold=True)
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Column widths
    ws.column_dimensions['A'].width = 35  # PARTICULARS
    ws.column_dimensions['B'].width = 14  # TOTAL
    for i in range(2, 14):  # C through N (months)
        ws.column_dimensions[get_column_letter(i + 1)].width = 14
    ws.column_dimensions['O'].width = 14  # Verification
    ws.column_dimensions['P'].width = 14  # Difference

    # Header rows
    row = 1
    ws[f'A{row}'] = "NAME OF ASSESSEE:"
    ws[f'B{row}'] = client_name
    ws[f'B{row}'].font = bold
    row += 1
    ws[f'A{row}'] = "ASSESSMENT YEAR:"
    ws[f'B{row}'] = assessment_year
    ws[f'B{row}'].font = bold
    row += 1
    row += 1  # blank row

    # Write each account section
    for summary in summaries:
        row = _write_account_section(ws, summary, row, bold, header_font, thin_border)
        row += 2  # gap between sections

    # Write detail sheets (one per account)
    if accounts:
        for account in accounts:
            _write_detail_sheet(wb, account, bold, header_font, thin_border)

    wb.save(output_path)
    return output_path


def _write_account_section(
    ws, summary: AccountSummary, start_row: int,
    bold: Font, header_font: Font, border: Border,
) -> int:
    """Write one bank account section. Returns the next available row."""
    row = start_row

    # Bank & Branch header
    ws[f'A{row}'] = f"BANK & BRANCH:"
    ws[f'B{row}'] = f"{summary.bank_name}  {summary.branch}"
    ws[f'B{row}'].font = bold
    row += 1

    ws[f'A{row}'] = "A/C NO.:"
    ws[f'B{row}'] = summary.account_number
    ws[f'B{row}'].font = bold
    row += 1

    ws[f'A{row}'] = "IFSC"
    ws[f'B{row}'] = summary.ifsc_code
    row += 1

    # Column headers: PARTICULARS | TOTAL | APR | ... | MAR
    header_row = row
    headers = ["PARTICULARS", "TOTAL"] + FY_MONTHS
    for i, h in enumerate(headers):
        cell = ws[f'{_col(i)}{row}']
        cell.value = h
        cell.font = header_font
        cell.border = border
    row += 1

    # Opening Balance
    ob_row = row
    ws[f'A{row}'] = "Opening Balance."
    ws[f'A{row}'].font = bold
    # B = same as opening balance value (TOTAL column = first month)
    ws[f'B{row}'] = float(summary.opening_balance)

    # C (APR) = opening balance
    ws[f'C{row}'] = f'=B{row}'
    # D through N: each month's opening = previous month's closing
    # (will be filled as formulas referencing closing balance row after we know it)
    # Placeholder - will be updated after closing balance row is known
    row += 1

    # "Deposits :-"
    ws[f'A{row}'] = "Deposits :-"
    ws[f'A{row}'].font = bold
    row += 1

    # Deposit category rows
    dep_start_row = row
    for cat in summary.deposit_categories:
        ws[f'A{row}'] = f"  {cat.category}"
        # B = SUM(C:N) formula
        ws[f'B{row}'] = f'=SUM(C{row}:N{row})'
        for mi in range(12):
            val = float(cat.monthly[mi])
            if val != 0:
                ws[f'{_col(mi + 2)}{row}'] = val
        row += 1
    dep_end_row = row - 1

    # Deposits Total Row
    dep_total_row = row
    ws[f'A{row}'] = "  Total Rs."
    ws[f'A{row}'].font = bold
    if dep_start_row <= dep_end_row:
        ws[f'B{row}'] = f'=SUM(B{dep_start_row}:B{dep_end_row})'
        for mi in range(12):
            col_letter = _col(mi + 2)
            ws[f'{col_letter}{row}'] = f'=SUM({col_letter}{dep_start_row}:{col_letter}{dep_end_row})'
    else:
        ws[f'B{row}'] = 0
        for mi in range(12):
            ws[f'{_col(mi + 2)}{row}'] = 0
    row += 1

    # "Withdrawls:-"
    ws[f'A{row}'] = "Withdrawls:-"
    ws[f'A{row}'].font = bold
    row += 1

    # Withdrawal category rows
    wdl_start_row = row
    for cat in summary.withdrawal_categories:
        ws[f'A{row}'] = f"  {cat.category}"
        ws[f'B{row}'] = f'=SUM(C{row}:N{row})'
        for mi in range(12):
            val = float(cat.monthly[mi])
            if val != 0:
                ws[f'{_col(mi + 2)}{row}'] = val
        row += 1
    wdl_end_row = row - 1

    # Withdrawals Total Row
    wdl_total_row = row
    ws[f'A{row}'] = "  Total Rs."
    ws[f'A{row}'].font = bold
    if wdl_start_row <= wdl_end_row:
        ws[f'B{row}'] = f'=SUM(B{wdl_start_row}:B{wdl_end_row})'
        for mi in range(12):
            col_letter = _col(mi + 2)
            ws[f'{col_letter}{row}'] = f'=SUM({col_letter}{wdl_start_row}:{col_letter}{wdl_end_row})'
    else:
        ws[f'B{row}'] = 0
        for mi in range(12):
            ws[f'{_col(mi + 2)}{row}'] = 0
    row += 1

    # Closing Balance
    closing_row = row
    ws[f'A{row}'] = "Closing Balance"
    ws[f'A{row}'].font = bold
    # B (TOTAL) = Opening + Deposits - Withdrawals
    ws[f'B{row}'] = f'=B{ob_row}+B{dep_total_row}-B{wdl_total_row}'
    # Per month: Closing = Opening + Deposits - Withdrawals
    for mi in range(12):
        col = _col(mi + 2)
        ws[f'{col}{row}'] = f'={col}{ob_row}+{col}{dep_total_row}-{col}{wdl_total_row}'
    row += 1

    # Now go back and fill in opening balance formulas for months D through N
    # D (MAY) opening = C (APR) closing, etc.
    for mi in range(1, 12):
        prev_col = _col(mi + 1)  # previous month column
        curr_col = _col(mi + 2)  # current month column
        ws[f'{curr_col}{ob_row}'] = f'={prev_col}{closing_row}'

    # Verification column O = last month's closing (independent check value)
    # P = N{closing} - O{closing} (should be 0)
    # We leave O blank (user fills it manually) and set P as formula
    ws[f'P{closing_row}'] = f'=N{closing_row}-O{closing_row}'

    return row


def _write_detail_sheet(
    wb: Workbook, account: BankAccount,
    bold: Font, header_font: Font, border: Border,
) -> None:
    """Write a detail sheet for one account with transactions grouped by category."""
    # Sheet name: truncate to 31 chars (Excel limit), ensure uniqueness
    acct_short = account.account_number[-6:] if len(account.account_number) > 6 else account.account_number
    sheet_name = f"{account.bank_name[:20]} {acct_short}"[:31]

    # Ensure unique sheet name
    existing = {ws.title for ws in wb.worksheets}
    if sheet_name in existing:
        sheet_name = sheet_name[:28] + f" {len(existing)}"

    ws = wb.create_sheet(title=sheet_name)

    # Column widths
    ws.column_dimensions['A'].width = 14  # Date
    ws.column_dimensions['B'].width = 50  # Description
    ws.column_dimensions['C'].width = 14  # Debit
    ws.column_dimensions['D'].width = 14  # Credit
    ws.column_dimensions['E'].width = 14  # Balance

    # Account header
    row = 1
    ws[f'A{row}'] = f"{account.bank_name} — A/C {account.account_number}"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1
    if account.holder_name:
        ws[f'A{row}'] = account.holder_name
        row += 1
    row += 1

    # Group transactions by category, preserving type (credit/debit)
    deposit_groups: dict[str, list] = defaultdict(list)
    withdrawal_groups: dict[str, list] = defaultdict(list)
    for txn in account.transactions:
        if txn.is_credit:
            deposit_groups[txn.category].append(txn)
        elif txn.is_debit:
            withdrawal_groups[txn.category].append(txn)

    cat_header_fill = PatternFill(start_color="E7EEF7", end_color="E7EEF7", fill_type="solid")
    dep_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    wdl_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    # --- Deposits ---
    if deposit_groups:
        ws[f'A{row}'] = "DEPOSITS"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="2B7A5B")
        row += 1

        for category, txns in sorted(deposit_groups.items(), key=lambda x: sum(t.credit for t in x[1]), reverse=True):
            # Category header
            cat_total = sum(t.credit for t in txns)
            ws[f'A{row}'] = category
            ws[f'A{row}'].font = bold
            ws[f'D{row}'] = f"{len(txns)} txns"
            ws[f'E{row}'] = float(cat_total)
            ws[f'E{row}'].font = bold
            for c in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{c}{row}'].fill = cat_header_fill
            row += 1

            # Column headers
            for i, h in enumerate(["Date", "Description", "Debit", "Credit", "Balance"]):
                cell = ws[f'{_col(i)}{row}']
                cell.value = h
                cell.font = Font(bold=True, size=9)
                cell.border = border
            row += 1

            # Transaction rows
            for txn in sorted(txns, key=lambda t: t.date):
                ws[f'A{row}'] = txn.date.strftime("%d/%m/%Y")
                ws[f'B{row}'] = txn.description
                if txn.debit:
                    ws[f'C{row}'] = float(txn.debit)
                if txn.credit:
                    ws[f'D{row}'] = float(txn.credit)
                if txn.balance is not None:
                    ws[f'E{row}'] = float(txn.balance)
                for c in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{c}{row}'].fill = dep_fill
                row += 1

            row += 1  # gap between categories

    # --- Withdrawals ---
    if withdrawal_groups:
        ws[f'A{row}'] = "WITHDRAWALS"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="C4653F")
        row += 1

        for category, txns in sorted(withdrawal_groups.items(), key=lambda x: sum(t.debit for t in x[1]), reverse=True):
            cat_total = sum(t.debit for t in txns)
            ws[f'A{row}'] = category
            ws[f'A{row}'].font = bold
            ws[f'C{row}'] = f"{len(txns)} txns"
            ws[f'E{row}'] = float(cat_total)
            ws[f'E{row}'].font = bold
            for c in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{c}{row}'].fill = cat_header_fill
            row += 1

            for i, h in enumerate(["Date", "Description", "Debit", "Credit", "Balance"]):
                cell = ws[f'{_col(i)}{row}']
                cell.value = h
                cell.font = Font(bold=True, size=9)
                cell.border = border
            row += 1

            for txn in sorted(txns, key=lambda t: t.date):
                ws[f'A{row}'] = txn.date.strftime("%d/%m/%Y")
                ws[f'B{row}'] = txn.description
                if txn.debit:
                    ws[f'C{row}'] = float(txn.debit)
                if txn.credit:
                    ws[f'D{row}'] = float(txn.credit)
                if txn.balance is not None:
                    ws[f'E{row}'] = float(txn.balance)
                for c in ['A', 'B', 'C', 'D', 'E']:
                    ws[f'{c}{row}'].fill = wdl_fill
                row += 1

            row += 1
